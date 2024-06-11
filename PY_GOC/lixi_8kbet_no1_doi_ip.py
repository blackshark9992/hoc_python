import os
import requests
import base64
import pytesseract
from PIL import Image
import uuid
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from concurrent.futures import ThreadPoolExecutor, as_completed
import random
import re

# Đọc URLs từ file
with open('link_8kbet.txt', 'r') as file:
    base_urls = [line.strip() for line in file.readlines()]

# Đặt đường dẫn tới tesseract.exe
pytesseract.pytesseract.tesseract_cmd = r'C:\Tesseract-OCR\tesseract.exe'

# Đường dẫn tới thư mục chứa các file Excel
directory = os.path.dirname(os.path.abspath(__file__))

# Tìm tất cả các file Excel trong thư mục
excel_files = [f for f in os.listdir(directory) if f.endswith('.xlsx') or f.endswith('.xls')]

if not excel_files:
    print("Không tìm thấy tệp Excel nào trong thư mục.")
    exit(1)
else:
    print("=========Lựa Chọn Binh Khí=============")
    for idx, file in enumerate(excel_files, 1):
        print(f"{idx}. {file}")

    choice = int(input("Ra Đòn Đi: "))
    selected_file = excel_files[choice - 1]
    excel_path = os.path.join(directory, selected_file)

    try:
        # Đọc file Excel
        df = pd.read_excel(excel_path, header=None)
    except (InvalidFileException, ValueError) as e:
        print("Không thể đọc tệp Excel. Có thể tệp bị hỏng hoặc không phải là tệp Excel hợp lệ.")
        print("Chi tiết lỗi:", e)
        exit(1)

    # Bộ lưu trữ các tài khoản đã xử lý
    processed_accounts = set()

    # Đọc proxy từ file
    with open('proxyt.txt', 'r') as file:
        proxy_data = [line.strip().split('|') for line in file.readlines()]
        proxys = [proxy for _, proxy in proxy_data]
        api_keys = [api_key for api_key, _ in proxy_data]

    initial_sleep_time = int(input("Nhập thời gian đổi ip: "))

    # Hàm lấy IP hiện tại qua proxy
    def get_current_ip(session):
        try:
            response = session.get("http://api.ipify.org/?format=json", timeout=5)
            if response.status_code == 200:
                ip_info = response.json()
                return ip_info.get("ip")
            else:
                return None
        except requests.RequestException:
            return None

    # Hàm trích xuất văn bản từ hình ảnh bằng Tesseract
    def demo_imagetotext(path: str):
        img = Image.open(path)
        text = pytesseract.image_to_string(img)
        return text.replace('/', '').replace("\n", '')[:4]
    def extract_wait_time(message):
        match = re.search(r'\d+', message)
        if match:
            return int(match.group())
        return None
    # Hàm thay đổi IP
    def change_ip(api_key):
        time.sleep(initial_sleep_time)
        url = f"https://app.proxyno1.com/api/change-key-ip/{api_key}"
        while True:
            try:
                response = requests.get(url)
                if response.status_code == 200:
                    data = response.json()
                    if data.get("status") == 0:
                        print("Đổi IP thành công")
                        time.sleep(10)
                        return True
                    else:
                        wait_time = extract_wait_time(data.get("message"))
                        if wait_time:
                            print(f"Chờ {wait_time}.")
                            time.sleep(wait_time)
                else:
                    print("Lỗi khi gọi API đổi IP :", response.status_code)
                    return False
            except requests.RequestException as e:
                print("Lỗi kết nối khi gọi API đổi IP:", e)
                return False

    # Hàm trích xuất thời gian chờ từ thông báo lỗi
    def extract_wait_time(message):
        match = re.search(r'\d+', message)
        if match:
            return int(match.group(0))
        return None

    # Hàm lấy captcha và đăng nhập
    def get_captcha_and_login(user, password, proxy, api_key):
        if not change_ip(api_key):
            return None, None, "Lỗi khi đổi IP"

        base_url = random.choice(base_urls)
        url_captcha = base_url + "/api/0.0/Home/GetCaptchaForLogin"
        api_login = base_url + "/api/0.0/Login/login"

        proxies = {
            "http": "http://" + proxy,
            "https": "http://" + proxy
        }
        session = requests.Session()
        session.proxies.update(proxies)

        # In IP mới
        retry_count = 0
        while retry_count < 3:
            new_ip = get_current_ip(session)
            if new_ip:
                pass
                break
            else:
                print("Không thể lấy IP qua proxy. Chờ 30s")
                time.sleep(30)
                retry_count += 1

        if retry_count == 3:
            print("Không thể lấy IP sau 3 lần thử.")
            return None, session, "Proxy không hoạt động"

        while True:
            try:
                response = session.post(url_captcha)
                if response.status_code == 200:
                    data = response.json()
                    image_base64 = data.get("image")
                    value_key = data.get("value")

                    if image_base64:
                        image_data = base64.b64decode(image_base64)
                        image_path = f"{uuid.uuid4()}.png"
                        with open(image_path, "wb") as file:
                            file.write(image_data)
                        absolute_image_path = os.path.abspath(image_path)
                        text = demo_imagetotext(absolute_image_path)
                        os.remove(absolute_image_path)

                        payload = {
                            "account": user,
                            "checkCode": text,
                            "checkCodeEncrypt": value_key,
                            "fingerprint": str(uuid.uuid4()),
                            "password": password
                        }

                        response_login = session.post(api_login, json=payload)
                        if response_login.status_code == 200:
                            login_data = response_login.json()
                            if login_data.get("IsSuccess"):
                                access_token = login_data.get("LoginToken").get("AccessToken")
                                return access_token, session, None
                            else:
                                error_message = login_data.get("ErrorMessage")
                                print("Đăng nhập không thành công:", error_message)
                                if error_message.strip() == "Tài khoản mật khẩu sai":
                                    return None, session, "Tài khoản hoặc mật khẩu sai"
                                elif error_message.strip() == ("Tài khoản này đang bị vô hiệu hóa,"
                                                               " vui lòng liên hệ với bộ phận chăm sóc khách hàng"):
                                    return None, session, "Tài khoản bị khoá"
                        else:
                            pass
                    else:
                        pass
                else:
                    print("Lỗi khi lấy captcha. Mã trạng thái:", response.status_code)
                time.sleep(1)
            except requests.RequestException as e:
                print("Lỗi kết nối tới proxy:", e)
                return None, session, "Lỗi kết nối tới proxy"

    # Hàm lấy số dư tài khoản
    def get_balance(session, headers):
        base_url = random.choice(base_urls)
        tien = base_url + "/api/0.0/Account/GetMyBalance"
        try:
            response = session.post(tien, headers=headers)
            if response.status_code == 200:
                return response.text
            else:
                print("Lỗi khi lấy số dư tài khoản:", response.status_code)
                return None
        except requests.RequestException as e:
            print("Lỗi kết nối tới proxy khi lấy số dư:", e)
            return None

    # Hàm lấy thông tin bổ sung và xử lý bao lì xì
    def get_additional_info(user, password, index, proxy, api_key, sheet):
        base_url = random.choice(base_urls)
        get_red_envelop_list = base_url + "/api/0.0/RedEnvelope/GetRedEnvelopList"
        red_envelope_received = base_url + "/api/1.0/redEnvelope/received"

        access_token, session, login_error = get_captcha_and_login(user, password, proxy, api_key)
        if login_error:
            sheet.cell(row=index + 1, column=6).value = login_error  # Cột F
            book.save(excel_path)
            return index, None, None, None, login_error
        if not access_token:
            return index, None, None, None, None

        headers = {
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/json"
        }

        balance_before = get_balance(session, headers)
        if balance_before:
            pass

        red_envelop_amount = 0
        try:
            response_get_red_envelop_list = session.post(get_red_envelop_list, headers=headers)
            if response_get_red_envelop_list.status_code == 200:
                red_envelop_list_data = response_get_red_envelop_list.json()
                if isinstance(red_envelop_list_data, list):
                    for red_envelop in red_envelop_list_data:
                        id_from_list = red_envelop.get("Id")
                        if id_from_list:
                            red_envelope_payload = {"id": id_from_list}
                            response_red_envelope = session.post(red_envelope_received, headers=headers,
                                                                 json=red_envelope_payload)
                            if response_red_envelope.status_code == 200:
                                content_length = red_envelop.get("Amount")
                                if content_length:
                                    red_envelop_amount += float(content_length)
                            else:
                                print("Lỗi khi gọi API redEnvelope/received:", response_red_envelope.status_code)
                else:
                    print("Dữ liệu trả về không phải là danh sách.")
            else:
                print("Lỗi khi gọi API GetRedEnvelopList:", response_get_red_envelop_list.status_code)
        except requests.RequestException as e:
            print("Lỗi kết nối tới proxy khi lấy danh sách bao lì xì:", e)

        balance_after = get_balance(session, headers)
        if balance_after:
            print(f"============================\nTên        : {user}\nMật Khẩu   : {password}\nTiền Trước : "
                  f"{balance_before}\nLixi:      : {red_envelop_amount}"
                  f"\nTiền Sau   : {balance_after}\n============================")

        # Ghi kết quả vào file Excel
        if balance_before is not None:
            sheet.cell(row=index + 1, column=3).value = float(balance_before)
        if balance_after is not None:
            sheet.cell(row=index + 1, column=4).value = float(balance_after)
        sheet.cell(row=index + 1, column=5).value = red_envelop_amount

        book.save(excel_path)
        return index, balance_before, balance_after, red_envelop_amount, None

    # Hàm xử lý từng dòng
    def process_row(index, row, proxy, api_key, sheet):
        user = row.iloc[0]
        password = row.iloc[1]
        if pd.isna(user) or pd.isna(password):
            return index, None, None, None, None
        if user in processed_accounts:  # Kiểm tra xem tài khoản đã được xử lý chưa
            return index, None, None, None, "Tài khoản đã được xử lý"
        processed_accounts.add(user)
        return get_additional_info(user, password, index, proxy, api_key, sheet)

    try:
        book = load_workbook(excel_path)
        sheet = book.active

        # Sử dụng ThreadPoolExecutor để chạy đa luồng dựa trên số lượng proxy
        num_proxies = len(proxys)
        with ThreadPoolExecutor(max_workers=num_proxies) as executor:
            futures = []
            for index, row in df.iterrows():
                proxy = proxys[index % num_proxies]  # Gán proxy theo vòng lặp
                api_key = api_keys[index % num_proxies]  # Gán API key theo vòng lặp
                futures.append(executor.submit(process_row, index, row, proxy, api_key, sheet))

            for future in as_completed(futures):
                index, balance_before, balance_after, red_envelop_amount, login_error = future.result()
                if balance_before is not None or balance_after is not None or red_envelop_amount is not None:
                    print(f"Đã Chạy: {index + 1}")
                if login_error:
                    print(f"Lỗi đăng nhập cho tài khoản dòng {index + 1}: {login_error}")
        book.close()
    except InvalidFileException as e:
        print("Không thể mở tệp Excel. Có thể tệp bị hỏng hoặc không phải là tệp Excel hợp lệ.")
        print("Chi tiết lỗi:", e)
