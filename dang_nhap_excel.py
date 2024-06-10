import os
import requests
import base64
import pytesseract
from PIL import Image
import uuid
import time
import pandas as pd
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed

# URLs cho API
url_captcha = "https://www.8kbbb.com/api/0.0/Home/GetCaptchaForLogin"
api_login = "https://www.8kbbb.com/api/0.0/Login/login"
tien = "https://m.8kbbb.com/api/0.0/Account/GetMyBalance"
get_red_envelop_list = "https://m.8kbbb.com/api/0.0/RedEnvelope/GetRedEnvelopList"
red_envelope_received = "https://m.8kbbb.com/api/1.0/redEnvelope/received"

# Đặt đường dẫn tới tesseract.exe
pytesseract.pytesseract.tesseract_cmd = r'C:\Tesseract-OCR\tesseract.exe'

# Đường dẫn tới thư mục chứa các file Excel
directory = os.path.dirname(os.path.abspath(__file__))

# Tìm tất cả các file Excel trong thư mục
excel_files = [f for f in os.listdir(directory) if f.endswith('.xlsx') or f.endswith('.xls')]

if not excel_files:
    print("Không tìm thấy tệp Excel nào trong thư mục.")
else:
    print("Các tệp Excel tìm thấy trong thư mục:")
    for idx, file in enumerate(excel_files, 1):
        print(f"{idx}. {file}")

    choice = int(input("Nhập số của tệp bạn muốn xử lý: "))
    selected_file = excel_files[choice - 1]
    excel_path = os.path.join(directory, selected_file)

    # Đọc file Excel
    df = pd.read_excel(excel_path, header=None)

    # Bộ lưu trữ các tài khoản đã xử lý
    processed_accounts = set()


    # Hàm trích xuất văn bản từ hình ảnh bằng Tesseract
    def demo_imagetotext(path: str):
        img = Image.open(path)
        text = pytesseract.image_to_string(img)
        return text.replace('/', '').replace("\n", '')[:4]


    # Hàm lấy captcha và đăng nhập
    def get_captcha_and_login(user, password):
        session = requests.Session()

        while True:
            response = session.post(url_captcha)
            if response.status_code == 200:
                data = response.json()
                image_base64 = data.get("image")
                value_key = data.get("value")

                if image_base64:
                    image_data = base64.b64decode(image_base64)
                    image_path = f"{uuid.uuid4()}.png"
                    with open(image_path, "wb") as file:
                        file.write(image_data)
                    absolute_image_path = os.path.abspath(image_path)
                    text = demo_imagetotext(absolute_image_path)
                    os.remove(absolute_image_path)

                    payload = {
                        "account": user,
                        "checkCode": text,
                        "checkCodeEncrypt": value_key,
                        "fingerprint": str(uuid.uuid4()),
                        "password": password
                    }

                    response_login = session.post(api_login, json=payload)
                    if response_login.status_code == 200:
                        login_data = response_login.json()
                        if login_data.get("IsSuccess"):
                            print(f"Tài Khoản: {user}|{password}")
                            access_token = login_data.get("LoginToken").get("AccessToken")
                            return access_token, session, None
                        else:
                            error_message = login_data.get("ErrorMessage")
                            print("Đăng nhập không thành công:", error_message)
                            if error_message.strip() == "Tài khoản mật khẩu sai":
                                return None, session, "Tài khoản hoặc mật khẩu sai"
                            elif error_message.strip() == ("Tài khoản này đang bị vô hiệu hóa,"
                                                           " vui lòng liên hệ với bộ phận chăm sóc khách hàng"):
                                return None, session, "Tài khoản bị khoá"
                    else:
                        print("Lỗi trong quá trình đăng nhập:", response_login.status_code)
                else:
                    print("Không Thấy hình ảnh.")
            else:
                print("Lỗi khi lấy captcha. Mã trạng thái:", response.status_code)
            time.sleep(1)


    # Hàm lấy số dư tài khoản
    def get_balance(session, headers):
        response = session.post(tien, headers=headers)
        if response.status_code == 200:
            return response.text
        else:
            print("Lỗi khi lấy số dư tài khoản:", response.status_code)
            return None


    # Hàm lấy thông tin bổ sung và xử lý bao lì xì
    def get_additional_info(user, password, index):
        access_token, session, login_error = get_captcha_and_login(user, password)
        if login_error:
            return index, None, None, None, login_error
        if not access_token:
            return index, None, None, None, None

        headers = {
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/json"
        }

        balance_before = get_balance(session, headers)
        if balance_before:
            print("Số Tiền Trước:", balance_before)

        red_envelop_amount = 0
        response_get_red_envelop_list = session.post(get_red_envelop_list, headers=headers)
        if response_get_red_envelop_list.status_code == 200:
            red_envelop_list_data = response_get_red_envelop_list.json()
            if isinstance(red_envelop_list_data, list):
                for red_envelop in red_envelop_list_data:
                    id_from_list = red_envelop.get("Id")
                    if id_from_list:
                        red_envelope_payload = {"id": id_from_list}
                        response_red_envelope = session.post(red_envelope_received, headers=headers,
                                                             json=red_envelope_payload)
                        if response_red_envelope.status_code == 200:
                            content_length = response_red_envelope.request.headers.get('Content-Length')
                            if content_length:
                                red_envelop_amount += int(content_length)
                        else:
                            print("Lỗi khi gọi API redEnvelope/received:", response_red_envelope.status_code)
            else:
                print("Dữ liệu trả về không phải là danh sách.")
        else:
            print("Lỗi khi gọi API GetRedEnvelopList:", response_get_red_envelop_list.status_code)

        balance_after = get_balance(session, headers)
        if balance_after:
            print("Số Tiền Sau:", balance_after)

        return index, balance_before, balance_after, red_envelop_amount, None


    # Hàm xử lý từng dòng
    def process_row(index, row):
        user = row.iloc[0]
        password = row.iloc[1]
        if pd.isna(user) or pd.isna(password):
            return index, None, None, None, None
        if user in processed_accounts:  # Kiểm tra xem tài khoản đã được xử lý chưa
            return index, None, None, None, "Tài khoản đã được xử lý"
        processed_accounts.add(user)
        return get_additional_info(user, password, index)


    # Đọc workbook hiện có
    book = load_workbook(excel_path)
    sheet = book.active

    start_row = 1

    # Sử dụng ThreadPoolExecutor để chạy đa luồng
    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = [executor.submit(process_row, index, row) for index, row in df.iterrows()]

        for future in as_completed(futures):
            index, balance_before, balance_after, red_envelop_amount, login_error = future.result()
            # Bỏ qua kết quả nếu user hoặc password trống
            if balance_before is None and balance_after is None and red_envelop_amount is None:
                if login_error:
                    sheet.cell(row=start_row + index, column=6).value = login_error  # Cột F
                    # Lưu các thay đổi vào file Excel ngay sau khi ghi lỗi
                    book.save(excel_path)
                continue
            # Cập nhật file Excel với kết quả
            if balance_before is not None:
                sheet.cell(row=start_row + index, column=3).value = float(balance_before)  # Cột C
            if balance_after is not None:
                sheet.cell(row=start_row + index, column=4).value = float(balance_after)  # Cột D
            sheet.cell(row=start_row + index, column=5).value = red_envelop_amount  # Cột E

            book.save(excel_path)
            print(
                f"Đã Chạy: {start_row + index}")

    book.save(excel_path)
    book.close()
