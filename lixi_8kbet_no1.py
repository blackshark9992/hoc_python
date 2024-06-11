import os
import requests
import base64
import pytesseract
from PIL import Image
import uuid
import time
import pandas as pd
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
import random

# Đọc URLs từ file
with open('link_8kbet.txt', 'r') as file:
    base_urls = [line.strip() for line in file.readlines()]

# Đặt đường dẫn tới tesseract.exe
pytesseract.pytesseract.tesseract_cmd = r'C:\Tesseract-OCR\tesseract.exe'

# Đường dẫn tới thư mục chứa các file Excel
directory = os.path.dirname(os.path.abspath(__file__))
# print(base_urls)
# Tìm tất cả các file Excel trong thư mục
excel_files = [f for f in os.listdir(directory) if f.endswith('.xlsx') or f.endswith('.xls')]

if not excel_files:
    print("Không tìm thấy tệp Excel nào trong thư mục.")
else:
    print("=========Lựa Chọn Binh Khí=============")
    for idx, file in enumerate(excel_files, 1):
        print(f"{idx}. {file}")

    choice = int(input("Ra Đòn Đi: "))
    selected_file = excel_files[choice - 1]
    excel_path = os.path.join(directory, selected_file)

    # Đọc file Excel
    df = pd.read_excel(excel_path, header=None)

    # Bộ lưu trữ các tài khoản đã xử lý
    processed_accounts = set()

    # Đọc proxy từ file
    with open('proxyt.txt', 'r') as file:
        proxys = [line.strip() for line in file.readlines()]

    # Hàm lấy IP hiện tại qua proxy
    def get_current_ip(session):
        try:
            response = session.get("http://api.ipify.org/?format=json", timeout=5)
            if response.status_code == 200:
                ip_info = response.json()
                return ip_info.get("ip")
            else:
                return None
        except requests.RequestException:
            return None

    # Hàm trích xuất văn bản từ hình ảnh bằng Tesseract
    def demo_imagetotext(path: str):
        img = Image.open(path)
        text = pytesseract.image_to_string(img)
        return text.replace('/', '').replace("\n", '')[:4]

    # Hàm lấy captcha và đăng nhập
    def get_captcha_and_login(user, password, proxy):
        base_url = random.choice(base_urls)
        # print(f"Game: {base_url}")  # Print base URL for each login
        url_captcha = base_url + "/api/0.0/Home/GetCaptchaForLogin"
        api_login = base_url + "/api/0.0/Login/login"

        proxies = {
            "http": "http://" + proxy,
            "https": "http://" + proxy
        }
        session = requests.Session()
        session.proxies.update(proxies)

        # In IP mới
        new_ip = get_current_ip(session)
        if new_ip:
            pass
        else:
            print("Không thể lấy IP qua proxy.")
            return None, session, "Proxy không hoạt động"

        while True:
            try:
                response = session.post(url_captcha)
                if response.status_code == 200:
                    data = response.json()
                    image_base64 = data.get("image")
                    value_key = data.get("value")

                    if image_base64:
                        image_data = base64.b64decode(image_base64)
                        image_path = f"{uuid.uuid4()}.png"
                        with open(image_path, "wb") as file:
                            file.write(image_data)
                        absolute_image_path = os.path.abspath(image_path)
                        text = demo_imagetotext(absolute_image_path)
                        os.remove(absolute_image_path)

                        payload = {
                            "account": user,
                            "checkCode": text,
                            "checkCodeEncrypt": value_key,
                            "fingerprint": str(uuid.uuid4()),
                            "password": password
                        }

                        response_login = session.post(api_login, json=payload)
                        if response_login.status_code == 200:
                            login_data = response_login.json()
                            if login_data.get("IsSuccess"):
                                # print(f"Tài Khoản: {user}|{password}")
                                access_token = login_data.get("LoginToken").get("AccessToken")
                                return access_token, session, None
                            else:
                                error_message = login_data.get("ErrorMessage")
                                if error_message.strip() == "Tài khoản mật khẩu sai":
                                    return None, session, "Tài khoản hoặc mật khẩu sai"
                                elif error_message.strip() == ("Tài khoản này đang bị vô hiệu hóa,"
                                                               " vui lòng liên hệ với bộ phận chăm sóc khách hàng"):
                                    return None, session, "Tài khoản bị khoá"
                        else:
                            print("Lỗi trong quá trình đăng nhập:", response_login.status_code)
                    else:
                        print("Không Thấy hình ảnh.")
                else:
                    print("Lỗi khi lấy captcha. Mã trạng thái:", response.status_code)
                time.sleep(1)
            except requests.RequestException as e:
                print("Lỗi kết nối tới proxy:", e)
                return None, session, "Lỗi kết nối tới proxy"

    # Hàm lấy số dư tài khoản
    def get_balance(session, headers):
        base_url = random.choice(base_urls)
        tien = base_url + "/api/0.0/Account/GetMyBalance"
        try:
            response = session.post(tien, headers=headers)
            if response.status_code == 200:
                return response.text
            else:
                print("Lỗi khi lấy số dư tài khoản:", response.status_code)
                return None
        except requests.RequestException as e:
            print("Lỗi kết nối tới proxy khi lấy số dư:", e)
            return None

    # Hàm lấy thông tin bổ sung và xử lý bao lì xì
    def get_additional_info(user, password, index, proxy):
        base_url = random.choice(base_urls)
        get_red_envelop_list = base_url + "/api/0.0/RedEnvelope/GetRedEnvelopList"
        red_envelope_received = base_url + "/api/1.0/redEnvelope/received"

        access_token, session, login_error = get_captcha_and_login(user, password, proxy)
        if login_error:
            return index, None, None, None, login_error
        if not access_token:
            return index, None, None, None, None

        headers = {
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/json"
        }

        balance_before = get_balance(session, headers)
        if balance_before:
            pass

        red_envelop_amount = 0
        try:
            response_get_red_envelop_list = session.post(get_red_envelop_list, headers=headers)
            if response_get_red_envelop_list.status_code == 200:
                red_envelop_list_data = response_get_red_envelop_list.json()
                if isinstance(red_envelop_list_data, list):
                    for red_envelop in red_envelop_list_data:
                        id_from_list = red_envelop.get("Id")
                        if id_from_list:
                            red_envelope_payload = {"id": id_from_list}
                            response_red_envelope = session.post(red_envelope_received, headers=headers,
                                                                 json=red_envelope_payload)
                            if response_red_envelope.status_code == 200:
                                content_length = red_envelop.get("Amount")
                                if content_length:
                                    red_envelop_amount += float(content_length)
                            else:
                                print("Lỗi khi gọi API redEnvelope/received:", response_red_envelope.status_code)
                else:
                    print("Dữ liệu trả về không phải là danh sách.")
            else:
                print("Lỗi khi gọi API GetRedEnvelopList:", response_get_red_envelop_list.status_code)
        except requests.RequestException as e:
            print("Lỗi kết nối tới proxy khi lấy danh sách bao lì xì:", e)

        balance_after = get_balance(session, headers)
        if balance_after:
            print(f"Tên        : {user}\nMật Khẩu   : {password}\nTiền Trước : {balance_before}\nLixi:      : {red_envelop_amount}"
                  f"\nTiền Sau   : {balance_after}\n============================")

        return index, balance_before, balance_after, red_envelop_amount, None

    # Hàm xử lý từng dòng
    def process_row(index, row, proxy):
        user = row.iloc[0]
        password = row.iloc[1]
        if pd.isna(user) or pd.isna(password):
            return index, None, None, None, None
        if user in processed_accounts:  # Kiểm tra xem tài khoản đã được xử lý chưa
            return index, None, None, None, "Tài khoản đã được xử lý"
        processed_accounts.add(user)
        return get_additional_info(user, password, index, proxy)

    # Đọc workbook hiện có
    book = load_workbook(excel_path)
    sheet = book.active

    start_row = 1

    # Prompt user for number of threads
    num_threads = int(input("Nhập số luồng bạn muốn sử dụng: "))

    # Sử dụng ThreadPoolExecutor để chạy đa luồng
    with ThreadPoolExecutor(max_workers=num_threads) as executor:
        futures = []
        for index, row in df.iterrows():
            proxy = proxys[index % len(proxys)]  # Gán proxy theo vòng lặp
            futures.append(executor.submit(process_row, index, row, proxy))

        for future in as_completed(futures):
            index, balance_before, balance_after, red_envelop_amount, login_error = future.result()
            # Bỏ qua kết quả nếu user hoặc password trống
            if balance_before is None and balance_after is None and red_envelop_amount is None:
                if login_error:
                    sheet.cell(row=start_row + index, column=6).value = login_error  # Cột F
                    # Lưu các thay đổi vào file Excel ngay sau khi ghi lỗi
                    book.save(excel_path)
                continue
            # Cập nhật file Excel với kết quả
            if balance_before is not None:
                sheet.cell(row=start_row + index, column=3).value = float(balance_before)  # Cột C
            if balance_after is not None:
                sheet.cell(row=start_row + index, column=4).value = float(balance_after)
            sheet.cell(row=start_row + index, column=5).value = red_envelop_amount

            book.save(excel_path)
            print(f"Đã Chạy: {start_row + index}")

    book.save(excel_path)
    book.close()
